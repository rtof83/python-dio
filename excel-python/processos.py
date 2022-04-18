import openpyxl
import os
from month import month

# função month retorna mês, ano, semestre atuais

# path tabela a
path_a = ''
# path tabela b (adição do mês corrente na string do nome da planilha)
path_b = '' +  month()[0] + ' - '+ month()[1] + '.xlsx'
# path tabela c
path_c = ''
a = openpyxl.load_workbook(filename=path_a)
b = openpyxl.load_workbook(filename=path_b)
c = openpyxl.load_workbook(filename=path_c)

# variáveis recebem nome da tab
origem = a['']
destino = ''
destino2 = c['']

tipoPort = input('[1] - Portarias Instauradas\n[2] - Portarias Solucionadas\n')
if (tipoPort == '1'): destino = b['PORT. INSTAURADAS']
if (tipoPort == '2'): destino = b['PORTARIAS SOLUCIONADAS']

numProc = input('\nNúmero do processo: ')
founded = False

#------------------------------------------------------------------------------

for i in range(1, origem.max_row + 1):
  if ((origem.cell(row=i, column=1).value) == numProc):
    founded = True
    for j in range(5, destino.max_row + 1):
      # RECUPERAR TIPO PORTARIA
      if ((destino.cell(row=j, column=1).value) == None):
        tipo = origem.cell(row=i, column=2).value

        # NR. DOC. <-- PROCEDIMENTO
        destino.cell(row=j, column=1, value = origem.cell(row=i, column=1).value)

        # TIPO PORTARIA ----------------------------------------------------------------
        if (tipo == 'SAD'): destino.cell(row=j, column=2, value = 'X')
        elif (tipo == 'RIP'): destino.cell(row=j, column=3, value = 'X')
        elif (tipo == 'PR'): destino.cell(row=j, column=4, value = 'X')
        elif (tipo == 'PCD'): destino.cell(row=j, column=5, value = 'X')
        elif (tipo == 'PQD'): destino.cell(row=j, column=6, value = 'X')
        elif (tipo == 'IPM'): destino.cell(row=j, column=7, value = 'X')
        elif (tipo == 'PAD'): destino.cell(row=j, column=8, value = 'X')
        elif (tipo == 'PADS'): destino.cell(row=j, column=9, value = 'X')
        elif (tipo == 'PAE'): destino.cell(row=j, column=10, value = 'X')
        elif (tipo == 'APF'): destino.cell(row=j, column=11, value = 'X')
        #--------------------------------------------------------------------------------------------------

        # DATA BOLETIM
        #destino.cell(row=j, column=12, value = origem.cell(row=i, column=12).value.strftime('%d/%m/%Y'))
        destino.cell(row=j, column=12, value = origem.cell(row=i, column=12).value)
        # BOLETIM
        if (tipoPort == '1'): destino.cell(row=j, column=13, value = origem.cell(row=i, column=11).value)
        # SÍNTESE DO FATO
        destino.cell(row=j, column=14, value = origem.cell(row=i, column=10).value)
        # ENCARREGADO
        destino.cell(row=j, column=16, value = origem.cell(row=i, column=6).value)
        # ENVOLVIDO
        destino.cell(row=j, column=18, value = origem.cell(row=i, column=5).value)

        # PORTARIAS SOLUCIONADASE -------------------------------------------------------------------------
        if (tipoPort == '2'):
          # DATA BOLETIM SOLUÇÃO
          #destino.cell(row=j, column=12, value = origem.cell(row=i, column=34).value.strftime('%d/%m/%Y'))
          destino.cell(row=j, column=12, value = origem.cell(row=i, column=34).value)
          # BOLETIM SOLUÇÃO
          destino.cell(row=j, column=13, value = origem.cell(row=i, column=33).value)
          # SOLUÇÃO (TRECHO RESOLVE)
          destino.cell(row=j, column=19, value = origem.cell(row=i, column=32).value)

          # -----------------------------------------------------------------------------------------------

          # RESOLUTIVIDADE
          for k in range(8, destino2.max_row + 1):
            if ((destino2.cell(row=k, column=1).value) == None):
              # NR. DOC. <-- PROCEDIMENTO
              destino2.cell(row=k, column=2, value = origem.cell(row=i, column=1).value)
              # TIPO PORTARIA
              destino2.cell(row=k, column=1, value = tipo)
              # DATA DO FATO
              #destino2.cell(row=k, column=3, value = origem.cell(row=i, column=7).value.strftime('%d/%m/%Y'))
              destino2.cell(row=k, column=3, value = origem.cell(row=i, column=7).value)
              # DATA SOLUÇÃO
              destino2.cell(row=k, column=4, value = origem.cell(row=i, column=37).value)

              c.save(path_c)
              break
        
        b.save(path_b)
        break

    # ----------------------------------------------------------------------------------------------------

    print('\nProcesso', tipo, origem.cell(row=i, column=1).value, 'inserido com sucesso.')

    if (tipoPort == '1'):
      input('\nPressione ENTER para sair e abrir a Planilha de Controle.')
    elif (tipoPort == '2'):
      input('\nPressione ENTER para sair e abrir as Planilhas de Controle e Resolutividade.')
      os.system('start ..\\path\\Planilha.xlsx')

    os.system('start ..\\"path\\Planilha.xlsx"')

if (founded == False):
  print('\nProcesso não encontrado.')
  input('\nPressione ENTER para sair.')
  